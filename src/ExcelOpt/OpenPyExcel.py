# -*- coding: utf-8 -*-


import os, sys

from openpyxl.workbook import Workbook
# ExcelWriter,里面封装好了对Excel的写操作
from openpyxl.writer.excel import ExcelWriter

# get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B
from openpyxl.cell import get_column_letter


def WriteExcel(tFileName, tDict):
    if len(tFileName) == 0:
        return
    # 新建一个workbook
    wb = Workbook()
    # 新建一个excelWriter
    ew = ExcelWriter(workbook=wb)
    # 第一个sheet是ws
    ws = wb.worksheets[0]
    # wb.add_sheet()
    # 设置ws的名称
    ws.title = "第一主播"
    # 设置文件输出路径与名称
    ws.cell("A1").value = '23'
    ws.cell("B1").value = '34'
    # 向某个单元格中写入数据
    lIndex = 1
    for itorTuple in tDict:
        lIndex += 1
        cellA = str.format("A%s"% lIndex)
        cellB = str.format("B%s"% lIndex)
        ws.cell(cellA).value = itorTuple[0]
        ws.cell(cellB).value = itorTuple[1]
    # 最后保存文件
    ew.save(filename=tFileName)


def AddSheetToExcel(tWorkHanle, tSheetName, tTitle, tDictValue):
    if tWorkHanle is None:
        return None

    lSheet = tWorkHanle.add_sheet(tSheetName)
    for itor in tDictValue:
        lSheet.cell


def SaveExcel(tWorkHanle, tFileName):
    try:
        ew = ExcelWriter(workbook=tWorkHanle)
        ew.save(filename=tFileName)
    except Exception, exinfo:
        print exinfo


if __name__ == "__main__":
    lDict = [(1, 1), (2, 2), (3, 3)]
    WriteExcel("test.xls", lDict)
    pass
